import openpyxl
import json
import datetime

SRC = r'D:\SynologyDrive\hotelbeoordelingen.xlsx'
LAND_DISPLAY = {'zwitserland': 'Zwitserland'}


def norm_land(name):
    return LAND_DISPLAY.get(name, name)


def norm_werk_prive(v):
    if not v:
        return None
    v = str(v).strip().lower()
    if v in ('werk', 'prive'):
        return v
    return None


RED_RGB = {'FFFFC7CE', 'FFFF0000'}
YELLOW_RGB = {'FFFFEB9C'}
GREEN_RGB = {'FFC6EFCE', 'FF00B050'}
GREEN_THEME = {9}
YELLOW_THEME = {5}


def rating_from_fill(cell):
    fill = cell.fill
    if not fill or fill.patternType is None:
        return None
    fg = fill.fgColor
    if fg.type == 'rgb':
        if fg.rgb in RED_RGB:
            return 1
        if fg.rgb in YELLOW_RGB:
            return 3
        if fg.rgb in GREEN_RGB:
            return 4
        return None
    if fg.type == 'theme':
        if fg.theme in GREEN_THEME:
            return 4
        if fg.theme in YELLOW_THEME:
            return 3
        return None
    return None


wb = openpyxl.load_workbook(SRC, data_only=True)
records = []
skipped_dates = []

for sheet_name in wb.sheetnames:
    if sheet_name == 'Template':
        continue
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=2):
        vals = [c.value for c in row]
        provincie, stad, hotelnaam = vals[0], vals[1], vals[2]
        datum, aantal, werk_prive, opmerkingen = vals[7], vals[8], vals[9], vals[10]
        if not hotelnaam:
            continue

        hygiene = rating_from_fill(row[3])
        badkamer = rating_from_fill(row[4])
        ontbijt = rating_from_fill(row[5])
        bed = rating_from_fill(row[6])

        begin_datum = None
        eind_datum = None
        extra_note = None
        if isinstance(datum, datetime.datetime):
            begin_datum = datum.date().isoformat()
        elif isinstance(datum, str) and datum.strip():
            extra_note = f'Datum: {datum.strip()}'
            skipped_dates.append((sheet_name, hotelnaam, datum))

        opmerk_parts = [p for p in [opmerkingen, extra_note] if p]
        opmerkingen_final = ' | '.join(opmerk_parts) if opmerk_parts else None

        rec = {
            'land': norm_land(sheet_name),
            'provincie': (str(provincie).strip() if provincie else None),
            'stad': (str(stad).strip() if stad else None),
            'hotelnaam': str(hotelnaam).strip(),
            'hygiene': hygiene,
            'badkamer': badkamer,
            'ontbijt': ontbijt,
            'bed': bed,
            'begin_datum': begin_datum,
            'eind_datum': eind_datum,
            'aantal_keer_geweest': aantal if isinstance(aantal, (int, float)) else None,
            'werk_prive': norm_werk_prive(werk_prive),
            'opmerkingen': opmerkingen_final,
        }
        records.append(rec)

out_path = r'C:\Users\jarmo\Desktop\Hotelbeoordeling app\scripts\hotels.json'
with open(out_path, 'w', encoding='utf-8') as f:
    json.dump(records, f, ensure_ascii=False, indent=2)

print(f'{len(records)} records written to {out_path}')
if skipped_dates:
    print('Rows with non-standard date text (kept as note in opmerkingen):')
    for s in skipped_dates:
        print(' ', s)
